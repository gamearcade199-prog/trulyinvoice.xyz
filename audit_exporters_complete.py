"""
Comprehensive Exporter Audit - Check all fields are included
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'backend'))

from app.services.accountant_excel_exporter import AccountantExcelExporter
from app.services.csv_exporter_v2 import ProfessionalCSVExporterV2
from openpyxl import load_workbook
import csv
from datetime import datetime

# Complete sample invoice with ALL possible fields
COMPLETE_INVOICE = {
    'id': '12345-67890',
    'invoice_number': 'INV-2025-001',
    'invoice_date': '2025-11-01',
    'due_date': '2025-12-01',
    'vendor_name': 'ABC Suppliers Private Limited',
    'vendor_gstin': '27AABCU9603R1ZM',
    'vendor_address': '123 Business Park, Mumbai',
    'vendor_state': 'Maharashtra',
    'vendor_phone': '+91-9876543210',
    'vendor_email': 'contact@abcsuppliers.com',
    'customer_name': 'XYZ Corporation Ltd',
    'customer_gstin': '29AABCT1234F2Z5',
    'customer_address': '456 Tech Tower, Bangalore',
    'customer_state': 'Karnataka',
    'customer_phone': '+91-9123456789',
    'customer_email': 'billing@xyzcorp.com',
    'payment_status': 'Partial',
    'paid_amount': 50000.00,
    'subtotal': 85000.00,
    'cgst': 7650.00,
    'sgst': 7650.00,
    'igst': 0.00,
    'total_amount': 100300.00,
    'discount': 5000.00,
    'shipping_charges': 2000.00,
    'notes': 'Urgent delivery required. Payment terms: 30 days net.',
    'payment_terms': 'Net 30 days from invoice date',
    'purchase_order_number': 'PO-2025-ABC-123',
    'created_at': '2025-11-01T10:30:00',
    'updated_at': '2025-11-02T15:45:00',
    'line_items': [
        {
            'description': 'Premium Widget Model A',
            'hsn_sac': '8471',
            'quantity': 100,
            'unit': 'Pcs',
            'rate': 500.00,
            'amount': 50000.00,
            'cgst_rate': 9.0,
            'cgst_amount': 4500.00,
            'sgst_rate': 9.0,
            'sgst_amount': 4500.00,
            'igst_rate': 0.0,
            'igst_amount': 0.00,
        },
        {
            'description': 'Professional Service Subscription',
            'hsn_sac': '998314',
            'quantity': 1,
            'unit': 'Year',
            'rate': 35000.00,
            'amount': 35000.00,
            'cgst_rate': 9.0,
            'cgst_amount': 3150.00,
            'sgst_rate': 9.0,
            'sgst_amount': 3150.00,
            'igst_rate': 0.0,
            'igst_amount': 0.00,
        }
    ]
}

CRITICAL_FIELDS = [
    'invoice_number', 'invoice_date', 'due_date',
    'vendor_name', 'vendor_gstin', 'vendor_address', 'vendor_state',
    'customer_name', 'customer_gstin', 'customer_address', 'customer_state',
    'subtotal', 'cgst', 'sgst', 'igst', 'total_amount',
    'discount', 'shipping_charges', 'payment_status', 'paid_amount', 'notes'
]

def check_excel_exporter():
    """Check AccountantExcelExporter for missing fields"""
    print("\n" + "="*80)
    print("TESTING: AccountantExcelExporter")
    print("="*80)
    
    exporter = AccountantExcelExporter()
    output_file = 'test_complete_invoice.xlsx'
    
    try:
        exporter.export_invoice(COMPLETE_INVOICE, output_file)
        print(f"✅ Excel file created: {output_file}")
        
        # Load and check the file
        wb = load_workbook(output_file)
        
        # Check Invoice Summary sheet
        if 'Invoice Summary' in wb.sheetnames:
            ws = wb['Invoice Summary']
            print("\n📊 Invoice Summary Sheet:")
            
            # Get headers from row 7
            headers = []
            for col in range(1, 20):
                value = ws.cell(row=7, column=col).value
                if value:
                    headers.append(value)
                else:
                    break
            
            print(f"   Headers found: {headers}")
            
            # Check for critical fields
            missing = []
            for field in ['Invoice No', 'Vendor Name', 'Vendor GSTIN', 'Customer Name', 
                         'Total Amount', 'Payment Status']:
                if field not in headers:
                    missing.append(field)
            
            if missing:
                print(f"   ❌ Missing headers: {missing}")
            else:
                print(f"   ✅ All critical headers present")
            
            # Check data row (row 8)
            data_row = {}
            for idx, header in enumerate(headers, 1):
                value = ws.cell(row=8, column=idx).value
                data_row[header] = value
                
            print(f"\n   Data Row Sample:")
            for header, value in list(data_row.items())[:5]:
                print(f"      {header}: {value}")
                
            # Specifically check vendor GSTIN
            if 'Vendor GSTIN' in data_row:
                gstin_value = data_row['Vendor GSTIN']
                if gstin_value == COMPLETE_INVOICE['vendor_gstin']:
                    print(f"   ✅ Vendor GSTIN correctly populated: {gstin_value}")
                else:
                    print(f"   ❌ Vendor GSTIN mismatch!")
                    print(f"      Expected: {COMPLETE_INVOICE['vendor_gstin']}")
                    print(f"      Got: {gstin_value}")
            else:
                print(f"   ❌ Vendor GSTIN column not found in summary!")
        
        # Check Line Items sheet
        if 'Line Items' in wb.sheetnames:
            ws = wb['Line Items']
            print("\n📊 Line Items Sheet:")
            
            headers = []
            for col in range(1, 20):
                value = ws.cell(row=1, column=col).value
                if value:
                    headers.append(value)
                else:
                    break
            
            print(f"   Headers: {headers}")
            line_items_count = 0
            for row in range(2, 10):
                if ws.cell(row=row, column=1).value:
                    line_items_count += 1
                else:
                    break
            print(f"   Line items found: {line_items_count}")
            
        # Check other sheets
        print(f"\n📋 All Sheets: {wb.sheetnames}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

def check_csv_exporter():
    """Check CSVExporterV2 for missing fields"""
    print("\n" + "="*80)
    print("TESTING: CSVExporterV2")
    print("="*80)
    
    exporter = ProfessionalCSVExporterV2()
    output_file = 'test_complete_invoice.csv'
    
    try:
        exporter.export_invoices([COMPLETE_INVOICE], output_file)
        print(f"✅ CSV file created: {output_file}")
        
        # Read and check the file
        with open(output_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames
            
            print(f"\n📊 CSV Headers ({len(headers)} total):")
            for i, header in enumerate(headers[:10], 1):
                print(f"   {i}. {header}")
            if len(headers) > 10:
                print(f"   ... and {len(headers) - 10} more")
            
            # Check for critical fields
            missing = []
            field_mapping = {
                'invoice_number': 'Invoice Number',
                'vendor_name': 'Vendor Name',
                'vendor_gstin': 'Vendor GSTIN',
                'customer_name': 'Customer Name',
                'customer_gstin': 'Customer GSTIN',
                'total_amount': 'Total Amount',
            }
            
            for field_key, display_name in field_mapping.items():
                # Check both possible formats
                found = False
                for h in headers:
                    if field_key.replace('_', ' ').lower() in h.lower() or display_name.lower() in h.lower():
                        found = True
                        break
                if not found:
                    missing.append(display_name)
            
            if missing:
                print(f"\n   ❌ Missing fields: {missing}")
            else:
                print(f"\n   ✅ All critical fields present")
            
            # Read first row of data
            f.seek(0)  # Reset to start
            reader = csv.DictReader(f)
            first_row = next(reader)
            
            print(f"\n   Data Row Sample:")
            for key, value in list(first_row.items())[:5]:
                print(f"      {key}: {value}")
            
            # Check vendor GSTIN specifically
            vendor_gstin_found = False
            for key, value in first_row.items():
                if 'vendor' in key.lower() and 'gstin' in key.lower():
                    if value == COMPLETE_INVOICE['vendor_gstin']:
                        print(f"   ✅ Vendor GSTIN correctly populated: {value}")
                        vendor_gstin_found = True
                    else:
                        print(f"   ❌ Vendor GSTIN mismatch!")
                        print(f"      Expected: {COMPLETE_INVOICE['vendor_gstin']}")
                        print(f"      Got: {value}")
                        vendor_gstin_found = True
                    break
            
            if not vendor_gstin_found:
                print(f"   ❌ Vendor GSTIN field not found in CSV!")
                print(f"   Available fields with 'vendor': {[k for k in first_row.keys() if 'vendor' in k.lower()]}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

def check_bulk_export():
    """Test bulk export with multiple invoices"""
    print("\n" + "="*80)
    print("TESTING: Bulk Export with Multiple Invoices")
    print("="*80)
    
    invoices = [COMPLETE_INVOICE]
    
    # Add a second invoice with different data
    invoice2 = COMPLETE_INVOICE.copy()
    invoice2['invoice_number'] = 'INV-2025-002'
    invoice2['vendor_name'] = 'DEF Traders'
    invoice2['vendor_gstin'] = '24AABCU9603R1ZX'
    invoice2['total_amount'] = 75000.00
    invoices.append(invoice2)
    
    exporter = AccountantExcelExporter()
    output_file = 'test_bulk_export.xlsx'
    
    try:
        exporter.export_invoices(invoices, output_file)
        print(f"✅ Bulk Excel file created: {output_file}")
        
        wb = load_workbook(output_file)
        ws = wb['Invoice Summary']
        
        print(f"\n📊 Checking multiple invoice data:")
        
        # Check both invoice rows
        for row_num in [8, 9]:  # First two data rows
            invoice_no = ws.cell(row=row_num, column=1).value
            vendor_gstin = ws.cell(row=row_num, column=5).value  # Column 5 is Vendor GSTIN
            
            print(f"   Row {row_num}: Invoice {invoice_no}, Vendor GSTIN: {vendor_gstin}")
            
            if not vendor_gstin or vendor_gstin == '':
                print(f"      ❌ Vendor GSTIN is empty!")
            else:
                print(f"      ✅ Vendor GSTIN populated")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

# Run all checks
if __name__ == "__main__":
    print("="*80)
    print("COMPREHENSIVE EXPORTER AUDIT")
    print("Checking all fields are correctly exported")
    print("="*80)
    
    check_excel_exporter()
    check_csv_exporter()
    check_bulk_export()
    
    print("\n" + "="*80)
    print("✅ AUDIT COMPLETE")
    print("="*80)
    print("\nGenerated files:")
    print("  - test_complete_invoice.xlsx")
    print("  - test_complete_invoice.csv")
    print("  - test_bulk_export.xlsx")
    print("\nPlease check these files manually to verify all fields are present.")
