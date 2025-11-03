"""
Deep inspection of Complete Data sheet
"""
from openpyxl import load_workbook

file = 'test_bulk_invoices.xlsx'
wb = load_workbook(file)

print("="*80)
print(f"INSPECTING: Complete Data Sheet")
print("="*80)

if 'Complete Data' in wb.sheetnames:
    ws = wb['Complete Data']
    
    print(f"\n📊 Headers (Row 1):")
    headers = []
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value:
            headers.append(value)
            print(f"   Col {col}: {value}")
    
    print(f"\n📋 Total columns: {len(headers)}")
    
    # Check if vendor_gstin is in headers
    if 'vendor_gstin' in headers:
        col_idx = headers.index('vendor_gstin') + 1
        print(f"\n✅ vendor_gstin found at column {col_idx}")
        
        # Check first data row
        value = ws.cell(row=2, column=col_idx).value
        print(f"   First row value: {value}")
        
        if value and value != '':
            print(f"   ✅ vendor_gstin has value!")
        else:
            print(f"   ❌ vendor_gstin is EMPTY!")
    else:
        print(f"\n❌ vendor_gstin NOT in headers!")
        print(f"   Columns with 'vendor': {[h for h in headers if 'vendor' in h.lower()]}")
        print(f"   Columns with 'gstin': {[h for h in headers if 'gstin' in h.lower()]}")

# Check Invoice Summary sheet row 7
print("\n" + "="*80)
print("INSPECTING: Invoice Summary Sheet (Row 7)")
print("="*80)

if 'Invoice Summary' in wb.sheetnames:
    ws = wb['Invoice Summary']
    
    print(f"\nRow 7 (should be headers):")
    headers = []
    for col in range(1, 15):
        value = ws.cell(row=7, column=col).value
        if value:
            headers.append(value)
            print(f"   Col {col}: {value}")
    
    # Check for Vendor GSTIN
    if 'Vendor GSTIN' in headers:
        col_idx = headers.index('Vendor GSTIN') + 1
        print(f"\n✅ 'Vendor GSTIN' found at column {col_idx}")
        
        # Check rows 8 and 9 (data rows)
        for row in [8, 9]:
            invoice_no = ws.cell(row=row, column=1).value
            gstin = ws.cell(row=row, column=col_idx).value
            
            if invoice_no:
                print(f"   Row {row}: Invoice {invoice_no}, GSTIN: {gstin}")
                if gstin and gstin != '':
                    print(f"      ✅ GSTIN populated")
                else:
                    print(f"      ❌ GSTIN EMPTY!")
    else:
        print(f"\n❌ 'Vendor GSTIN' NOT in row 7!")

wb.close()
