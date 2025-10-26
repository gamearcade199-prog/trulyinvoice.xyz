"""
🚀 BULLETPROOF 10/10 EXCEL EXPORTER - GEMINI COMPATIBLE
========================================================

Uses ALL data from our bulletproof Gemini extractor including:
- 50+ fields from bulletproof extractor
- 4 comprehensive sheets
- Enhanced formatting and professional styling
- Complete data utilization (banking, customer, GST compliance)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any


class BulletproofExcelExporter:
    """
    Creates comprehensive Excel exports using ALL data from bulletproof Gemini extractor
    """
    
    def __init__(self):
        # Professional color scheme
        self.colors = {
            'primary': '2E86AB',      # Professional blue
            'secondary': '404040',     # Dark gray
            'accent': 'A23B72',       # Accent purple
            'success': '4CAF50',      # Green
            'warning': 'FF9800',      # Orange
            'light': 'F5F5F5',       # Light gray
            'white': 'FFFFFF'
        }
        
        # Font styles
        self.fonts = {
            'title': Font(name='Calibri', size=16, bold=True, color='FFFFFF'),
            'section': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
            'subsection': Font(name='Calibri', size=12, bold=True, color=self.colors['primary']),
            'label': Font(name='Calibri', size=10, bold=True),
            'normal': Font(name='Calibri', size=10),
            'total': Font(name='Calibri', size=12, bold=True),
            'currency': Font(name='Calibri', size=10, color=self.colors['success'])
        }
        
        # Border styles
        self.borders = {
            'thin': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            ),
            'double': Border(
                left=Side(style='double'),
                right=Side(style='double'),
                top=Side(style='double'),
                bottom=Side(style='double')
            )
        }
    
    def export_invoice(self, invoice_data: Dict, filename: str = None) -> str:
        """
        Export single invoice using ALL bulletproof Gemini data
        
        Args:
            invoice_data: Complete invoice data from bulletproof extractor
            filename: Optional custom filename
            
        Returns:
            Path to created Excel file
        """
        if not filename:
            invoice_num = invoice_data.get('invoice_number', 'INVOICE').replace('/', '_')
            vendor_name = invoice_data.get('vendor_name', 'VENDOR').replace(' ', '_')[:15]
            filename = f"Bulletproof_Invoice_{invoice_num}_{vendor_name}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        # Create workbook with 4 comprehensive sheets
        wb = Workbook()
        
        # Sheet 1: Enhanced Invoice Summary
        ws_invoice = wb.active
        ws_invoice.title = "Invoice Details"
        self._build_invoice_sheet(ws_invoice, invoice_data)
        
        # Sheet 2: Financial Breakdown
        ws_financial = wb.create_sheet("Financial Analysis")
        self._build_financial_sheet(ws_financial, invoice_data)
        
        # Sheet 3: GST Compliance
        ws_compliance = wb.create_sheet("GST Compliance")
        self._build_compliance_sheet(ws_compliance, invoice_data)
        
        # Sheet 4: Complete Raw Data
        ws_raw = wb.create_sheet("Raw Data")
        self._build_raw_data_sheet(ws_raw, invoice_data)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Bulletproof Excel export created: {filename}")
        return filename
    
    def export_multiple_invoices(self, invoices_data: List[Dict], filename: str = None) -> str:
        """
        Export multiple invoices using bulletproof format
        
        Args:
            invoices_data: List of invoice data dictionaries
            filename: Optional custom filename
            
        Returns:
            Path to created Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M')
            filename = f"Bulletproof_Multiple_Invoices_{len(invoices_data)}_records_{timestamp}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Summary sheet with all invoices
        ws_summary = wb.active
        ws_summary.title = "Invoices Summary"
        self._build_summary_sheet(ws_summary, invoices_data)
        
        # Detailed financial analysis
        ws_analysis = wb.create_sheet("Financial Analysis")
        self._build_bulk_analysis_sheet(ws_analysis, invoices_data)
        
        # Save workbook
        wb.save(filename)
        print(f"✅ Bulletproof bulk Excel export created: {filename} ({len(invoices_data)} invoices)")
        return filename
    
    def _build_invoice_sheet(self, ws, data: Dict):
        """Build enhanced invoice details sheet"""
        
        row = 1
        
        # Header with invoice number
        invoice_num = data.get('invoice_number', 'N/A')
        vendor_name = data.get('vendor_name', 'Unknown Vendor')
        
        self._add_section_header(ws, row, 'A', 'H', f"INVOICE {invoice_num} - {vendor_name}", self.colors['primary'])
        row += 2
        
        # Core invoice information (2-column layout)
        core_fields = [
            ('Invoice Number', data.get('invoice_number', 'N/A')),
            ('Invoice Date', data.get('invoice_date', 'N/A')),
            ('Due Date', data.get('due_date', 'N/A')),
            ('Currency', data.get('currency', 'INR')),
            ('Total Amount', f"₹{float(data.get('total_amount', 0)):,.2f}"),
            ('Payment Status', data.get('payment_status', 'Unknown'))
        ]
        
        for i, (label, value) in enumerate(core_fields):
            current_row = row + i
            ws[f'A{current_row}'] = label
            ws[f'A{current_row}'].font = self.fonts['label']
            ws[f'B{current_row}'] = value
            ws[f'B{current_row}'].font = self.fonts['normal']
            
            # Highlight total amount
            if 'Amount' in label:
                ws[f'B{current_row}'].font = self.fonts['currency']
        
        row += len(core_fields) + 2
        
        # Enhanced vendor information
        self._add_section_header(ws, row, 'A', 'H', "VENDOR INFORMATION", self.colors['secondary'])
        row += 1
        
        vendor_fields = [
            ('Vendor Name', data.get('vendor_name', 'N/A')),
            ('GSTIN', data.get('vendor_gstin', 'N/A')),
            ('PAN', data.get('vendor_pan', 'N/A')),
            ('Email', data.get('vendor_email', 'N/A')),
            ('Phone', data.get('vendor_phone', 'N/A')),
            ('Address', data.get('vendor_address', 'N/A')),
            ('State', data.get('vendor_state', 'N/A')),
            ('Pincode', data.get('vendor_pincode', 'N/A'))
        ]
        
        for label, value in vendor_fields:
            if value and value != 'N/A':
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.fonts['label']
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = self.fonts['normal']
                row += 1
        
        row += 1
        
        # Customer information (if available)
        if data.get('customer_name'):
            self._add_section_header(ws, row, 'A', 'H', "CUSTOMER INFORMATION", self.colors['secondary'])
            row += 1
            
            customer_fields = [
                ('Customer Name', data.get('customer_name', 'N/A')),
                ('Customer GSTIN', data.get('customer_gstin', 'N/A')),
                ('Customer Address', data.get('customer_address', 'N/A')),
                ('Customer State', data.get('customer_state', 'N/A')),
                ('Customer Phone', data.get('customer_phone', 'N/A'))
            ]
            
            for label, value in customer_fields:
                if value and value != 'N/A':
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = self.fonts['label']
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].font = self.fonts['normal']
                    row += 1
            
            row += 1
        
        # Banking details (if available)
        if data.get('bank_name') or data.get('account_number'):
            self._add_section_header(ws, row, 'A', 'H', "BANKING DETAILS", self.colors['accent'])
            row += 1
            
            banking_fields = [
                ('Bank Name', data.get('bank_name', 'N/A')),
                ('Account Number', data.get('account_number', 'N/A')),
                ('IFSC Code', data.get('ifsc_code', 'N/A')),
                ('SWIFT Code', data.get('swift_code', 'N/A'))
            ]
            
            for label, value in banking_fields:
                if value and value != 'N/A':
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = self.fonts['label']
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].font = self.fonts['normal']
                    row += 1
        
        # Auto-adjust column widths
        self._auto_adjust_columns(ws)
    
    def _build_financial_sheet(self, ws, data: Dict):
        """Build comprehensive financial breakdown sheet"""
        
        row = 1
        
        # Header
        self._add_section_header(ws, row, 'A', 'F', "FINANCIAL BREAKDOWN & ANALYSIS", self.colors['primary'])
        row += 2
        
        # Tax analysis
        self._add_section_header(ws, row, 'A', 'F', "TAX ANALYSIS", self.colors['secondary'])
        row += 1
        
        tax_fields = [
            ('Taxable Amount', data.get('taxable_amount', 0)),
            ('CGST Amount', data.get('cgst', 0)),
            ('SGST Amount', data.get('sgst', 0)),
            ('IGST Amount', data.get('igst', 0)),
            ('UGST Amount', data.get('ugst', 0)),
            ('CESS Amount', data.get('cess', 0)),
            ('Total GST', data.get('total_gst', 0)),
            ('VAT (Legacy)', data.get('vat', 0)),
            ('Service Tax (Legacy)', data.get('service_tax', 0)),
            ('TDS Amount', data.get('tds_amount', 0)),
            ('TDS Percentage', f"{data.get('tds_percentage', 0)}%" if data.get('tds_percentage') else 0),
            ('TCS Amount', data.get('tcs_amount', 0))
        ]
        
        for label, value in tax_fields:
            if value and (isinstance(value, str) or float(value) != 0):
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.fonts['label']
                if isinstance(value, str) and '%' in str(value):
                    ws[f'B{row}'] = value
                else:
                    ws[f'B{row}'] = f"₹{float(value):,.2f}" if value else "₹0.00"
                ws[f'B{row}'].font = self.fonts['normal']
                row += 1
        
        row += 2
        
        # Charges breakdown
        self._add_section_header(ws, row, 'A', 'F', "ADDITIONAL CHARGES", self.colors['secondary'])
        row += 1
        
        charges_fields = [
            ('Discount', data.get('discount', 0)),
            ('Discount Percentage', f"{data.get('discount_percentage', 0)}%" if data.get('discount_percentage') else 0),
            ('Shipping Charges', data.get('shipping_charges', 0)),
            ('Packing Charges', data.get('packing_charges', 0)),
            ('Handling Charges', data.get('handling_charges', 0)),
            ('Insurance Charges', data.get('insurance_charges', 0)),
            ('Other Charges', data.get('other_charges', 0)),
            ('Round Off', data.get('roundoff', 0))
        ]
        
        for label, value in charges_fields:
            if value and (isinstance(value, str) or float(value) != 0):
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.fonts['label']
                if isinstance(value, str) and '%' in str(value):
                    ws[f'B{row}'] = value
                else:
                    ws[f'B{row}'] = f"₹{float(value):,.2f}" if value else "₹0.00"
                ws[f'B{row}'].font = self.fonts['normal']
                row += 1
        
        self._auto_adjust_columns(ws)
    
    def _build_compliance_sheet(self, ws, data: Dict):
        """Build GST compliance and regulatory information"""
        
        row = 1
        
        # Header
        self._add_section_header(ws, row, 'A', 'F', "GST COMPLIANCE & REGULATORY DATA", self.colors['primary'])
        row += 2
        
        # GST Information
        self._add_section_header(ws, row, 'A', 'F', "GST DETAILS", self.colors['secondary'])
        row += 1
        
        gst_fields = [
            ('HSN Code', data.get('hsn_code', 'N/A')),
            ('SAC Code', data.get('sac_code', 'N/A')),
            ('Place of Supply', data.get('place_of_supply', 'N/A')),
            ('Invoice Type', data.get('invoice_type', 'N/A')),
            ('Supply Type', data.get('supply_type', 'N/A')),
            ('Reverse Charge', 'Yes' if data.get('reverse_charge') else 'No')
        ]
        
        for label, value in gst_fields:
            if value and value != 'N/A':
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = self.fonts['label']
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = self.fonts['normal']
                row += 1
        
        row += 2
        
        # Import/Export (if applicable)
        if data.get('bill_of_entry') or data.get('port_code'):
            self._add_section_header(ws, row, 'A', 'F', "IMPORT/EXPORT DETAILS", self.colors['secondary'])
            row += 1
            
            import_fields = [
                ('Bill of Entry', data.get('bill_of_entry', 'N/A')),
                ('Bill of Entry Date', data.get('bill_of_entry_date', 'N/A')),
                ('Port Code', data.get('port_code', 'N/A'))
            ]
            
            for label, value in import_fields:
                if value and value != 'N/A':
                    ws[f'A{row}'] = label
                    ws[f'A{row}'].font = self.fonts['label']
                    ws[f'B{row}'] = value
                    ws[f'B{row}'].font = self.fonts['normal']
                    row += 1
        
        self._auto_adjust_columns(ws)
    
    def _build_raw_data_sheet(self, ws, data: Dict):
        """Build complete raw data sheet with all extracted fields"""
        
        row = 1
        
        # Header
        self._add_section_header(ws, row, 'A', 'C', "COMPLETE EXTRACTED DATA", self.colors['primary'])
        row += 2
        
        # Table headers
        ws[f'A{row}'] = "Field Name"
        ws[f'B{row}'] = "Value"
        ws[f'C{row}'] = "Data Type"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = self.fonts['subsection']
            ws[f'{col}{row}'].fill = PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid')
        row += 1
        
        # All extracted data
        for key, value in sorted(data.items()):
            if not key.startswith('_'):  # Skip internal metadata
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value) if value is not None else 'N/A'
                ws[f'C{row}'] = type(value).__name__
                
                # Color code by data type
                if isinstance(value, (int, float)):
                    ws[f'C{row}'].fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                elif isinstance(value, bool):
                    ws[f'C{row}'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
                
                row += 1
        
        # Extraction metadata
        row += 2
        self._add_section_header(ws, row, 'A', 'C', "EXTRACTION METADATA", self.colors['accent'])
        row += 1
        
        metadata = data.get('_extraction_metadata', {})
        for key, value in metadata.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = str(value)
            ws[f'C{row}'] = "Metadata"
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _build_summary_sheet(self, ws, invoices_data: List[Dict]):
        """Build summary sheet for multiple invoices"""
        
        row = 1
        
        # Header
        self._add_section_header(ws, row, 'A', 'M', f"INVOICES SUMMARY ({len(invoices_data)} Invoices)", self.colors['primary'])
        row += 2
        
        # Table headers
        headers = [
            'Invoice Number', 'Date', 'Vendor Name', 'Vendor GSTIN', 'Customer Name',
            'Total Amount', 'Currency', 'CGST', 'SGST', 'IGST', 'Total GST',
            'Payment Status', 'Created At'
        ]
        
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=i, value=header)
            cell.font = self.fonts['subsection']
            cell.fill = PatternFill(start_color=self.colors['light'], end_color=self.colors['light'], fill_type='solid')
        row += 1
        
        # Data rows
        for invoice in invoices_data:
            data_row = [
                invoice.get('invoice_number', ''),
                invoice.get('invoice_date', ''),
                invoice.get('vendor_name', ''),
                invoice.get('vendor_gstin', ''),
                invoice.get('customer_name', ''),
                float(invoice.get('total_amount', 0)),
                invoice.get('currency', 'INR'),
                float(invoice.get('cgst', 0)),
                float(invoice.get('sgst', 0)),
                float(invoice.get('igst', 0)),
                float(invoice.get('total_gst', 0)),
                invoice.get('payment_status', ''),
                invoice.get('created_at', '')
            ]
            
            for i, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=i, value=value)
                cell.font = self.fonts['normal']
                
                # Format currency columns
                if i in [6, 8, 9, 10, 11] and isinstance(value, (int, float)):
                    cell.value = f"₹{value:,.2f}"
                    cell.font = self.fonts['currency']
                
                cell.border = self.borders['thin']
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _build_bulk_analysis_sheet(self, ws, invoices_data: List[Dict]):
        """Build financial analysis for bulk invoices"""
        
        row = 1
        
        # Header
        self._add_section_header(ws, row, 'A', 'F', "BULK FINANCIAL ANALYSIS", self.colors['primary'])
        row += 2
        
        # Calculate totals
        total_amount = sum(float(inv.get('total_amount', 0)) for inv in invoices_data)
        total_gst = sum(float(inv.get('total_gst', 0)) for inv in invoices_data)
        total_cgst = sum(float(inv.get('cgst', 0)) for inv in invoices_data)
        total_sgst = sum(float(inv.get('sgst', 0)) for inv in invoices_data)
        total_igst = sum(float(inv.get('igst', 0)) for inv in invoices_data)
        
        # Summary statistics
        summary_fields = [
            ('Total Invoices', len(invoices_data)),
            ('Total Invoice Amount', f"₹{total_amount:,.2f}"),
            ('Total GST Collected', f"₹{total_gst:,.2f}"),
            ('Total CGST', f"₹{total_cgst:,.2f}"),
            ('Total SGST', f"₹{total_sgst:,.2f}"),
            ('Total IGST', f"₹{total_igst:,.2f}"),
            ('Average Invoice Value', f"₹{total_amount/len(invoices_data):,.2f}" if invoices_data else "₹0.00"),
            ('Average GST per Invoice', f"₹{total_gst/len(invoices_data):,.2f}" if invoices_data else "₹0.00")
        ]
        
        for label, value in summary_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.fonts['label']
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.fonts['normal']
            if '₹' in str(value):
                ws[f'B{row}'].font = self.fonts['currency']
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _add_section_header(self, ws, row: int, start_col: str, end_col: str, title: str, color: str):
        """Add a formatted section header"""
        ws.merge_cells(f'{start_col}{row}:{end_col}{row}')
        cell = ws[f'{start_col}{row}']
        cell.value = title
        cell.font = self.fonts['section']
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
    
    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width