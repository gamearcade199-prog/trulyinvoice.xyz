"""
🏆 PROFESSIONAL EXCEL EXPORTER - Industry Standard
===================================================

Creates beautifully formatted Excel invoices with:
- Company branding header
- Proper sections (Vendor, Invoice Details, Line Items, Tax Summary)
- Color coding and formatting
- Professional styling (borders, bold, alignment)
- Multi-sheet support (Summary + Details)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, List, Any


class ProfessionalInvoiceExporter:
    """
    Creates professional Excel invoices matching industry standards
    """
    
    def __init__(self):
        # Color scheme (professional blue/grey)
        self.colors = {
            'header': '1F4E78',      # Dark blue
            'subheader': '5B9BD5',   # Light blue
            'total': 'F2F2F2',       # Light grey
            'white': 'FFFFFF',
            'warning': 'FFC7CE',     # Light red
            'success': 'C6EFCE'      # Light green
        }
        
        # Fonts
        self.fonts = {
            'title': Font(name='Calibri', size=18, bold=True, color='FFFFFF'),
            'header': Font(name='Calibri', size=12, bold=True, color='FFFFFF'),
            'subheader': Font(name='Calibri', size=11, bold=True),
            'normal': Font(name='Calibri', size=10),
            'bold': Font(name='Calibri', size=10, bold=True),
            'total': Font(name='Calibri', size=11, bold=True)
        }
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.borders = {
            'all': thin_border,
            'bottom': Border(bottom=Side(style='medium'))
        }
    
    def export_invoice(self, invoice_data: Dict, filename: str = None) -> str:
        """
        Export invoice to professional Excel format
        
        Args:
            invoice_data: Invoice data dictionary
            filename: Optional custom filename
            
        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        
        # Create sheets
        ws_invoice = wb.active
        ws_invoice.title = "Invoice"
        ws_details = wb.create_sheet("Details & Breakdown")
        
        # Build invoice sheet
        self._build_invoice_sheet(ws_invoice, invoice_data)
        
        # Build details sheet
        self._build_details_sheet(ws_details, invoice_data)
        
        # Save workbook
        if not filename:
            invoice_num = invoice_data.get('invoice_number', 'INVOICE').replace('/', '_')
            filename = f"Invoice_{invoice_num}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        wb.save(filename)
        print(f"✅ Professional invoice exported: {filename}")
        return filename
    
    def _build_invoice_sheet(self, ws, data: Dict):
        """Build main invoice sheet with professional formatting"""
        
        row = 1
        
        # ============ HEADER SECTION ============
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "INVOICE"
        cell.font = self.fonts['title']
        cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 1
        
        # Company info (if available)
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "TrulyInvoice.in - Professional Invoice Management"
        cell.font = Font(name='Calibri', size=10, italic=True)
        cell.alignment = Alignment(horizontal='center')
        row += 2
        
        # ============ VENDOR SECTION ============
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "VENDOR INFORMATION"
        cell.font = self.fonts['header']
        cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Vendor details
        vendor_fields = [
            ('Vendor Name', data.get('vendor_name', 'N/A')),
            ('GSTIN', data.get('vendor_gstin', 'N/A')),
            ('Address', data.get('vendor_address', 'N/A'))
        ]
        
        for label, value in vendor_fields:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = self.fonts['bold']
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = self.fonts['normal']
            row += 1
        
        row += 1
        
        # ============ INVOICE DETAILS SECTION ============
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "INVOICE DETAILS"
        cell.font = self.fonts['header']
        cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Split into two columns
        col1_fields = [
            ('Invoice Number', data.get('invoice_number', 'N/A')),
            ('Invoice Date', data.get('invoice_date', 'N/A')),
            ('Due Date', data.get('due_date', 'N/A'))
        ]
        
        col2_fields = [
            ('Payment Status', data.get('payment_status', 'unpaid').upper()),
            ('Currency', data.get('currency', 'INR')),
            ('Created At', data.get('created_at', 'N/A')[:10] if data.get('created_at') else 'N/A')
        ]
        
        start_row = row
        for i, (label, value) in enumerate(col1_fields):
            ws[f'A{row + i}'] = label
            ws[f'A{row + i}'].font = self.fonts['bold']
            ws[f'B{row + i}'] = value
            ws[f'B{row + i}'].font = self.fonts['normal']
        
        for i, (label, value) in enumerate(col2_fields):
            ws[f'D{row + i}'] = label
            ws[f'D{row + i}'].font = self.fonts['bold']
            ws[f'E{row + i}'] = value
            ws[f'E{row + i}'].font = self.fonts['normal']
            
            # Color code payment status
            if label == 'Payment Status':
                if 'PAID' in str(value).upper():
                    ws[f'E{row + i}'].fill = PatternFill(start_color=self.colors['success'], end_color=self.colors['success'], fill_type='solid')
                else:
                    ws[f'E{row + i}'].fill = PatternFill(start_color=self.colors['warning'], end_color=self.colors['warning'], fill_type='solid')
        
        row += max(len(col1_fields), len(col2_fields)) + 1
        
        # ============ LINE ITEMS SECTION ============
        if data.get('line_items') and len(data['line_items']) > 0:
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = "LINE ITEMS"
            cell.font = self.fonts['header']
            cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
            cell.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Table headers
            headers = ['#', 'Description', 'HSN/SAC', 'Qty', 'Rate', 'Amount']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = self.fonts['subheader']
                cell.fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')
                cell.border = self.borders['all']
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Line items
            for idx, item in enumerate(data['line_items'], start=1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = item.get('description', 'N/A')
                ws[f'C{row}'] = item.get('hsn_sac', item.get('hsn', 'N/A'))
                ws[f'D{row}'] = item.get('quantity', 1)
                ws[f'E{row}'] = f"₹{float(item.get('rate', 0)):,.2f}"
                ws[f'F{row}'] = f"₹{float(item.get('amount', 0)):,.2f}"
                
                # Apply borders to all cells in row
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = self.borders['all']
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center' if col in [1, 3, 4] else 'left')
                
                row += 1
            
            row += 1
        
        # ============ TAX SUMMARY SECTION ============
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "TAX SUMMARY & TOTALS"
        cell.font = self.fonts['header']
        cell.fill = PatternFill(start_color=self.colors['subheader'], end_color=self.colors['subheader'], fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Amounts (right-aligned)
        summary_items = [
            ('Subtotal', data.get('subtotal', 0)),
            ('CGST', data.get('cgst', 0)),
            ('SGST', data.get('sgst', 0)),
            ('IGST', data.get('igst', 0)),
        ]
        
        for label, value in summary_items:
            if value > 0 or label == 'Subtotal':  # Show subtotal even if 0
                ws[f'D{row}'] = label
                ws[f'D{row}'].font = self.fonts['normal']
                ws[f'D{row}'].alignment = Alignment(horizontal='right')
                
                ws[f'E{row}'] = f"₹{float(value):,.2f}"
                ws[f'E{row}'].font = self.fonts['normal']
                ws[f'E{row}'].alignment = Alignment(horizontal='right')
                row += 1
        
        # Total (highlighted)
        ws[f'D{row}'] = "TOTAL AMOUNT"
        ws[f'D{row}'].font = self.fonts['total']
        ws[f'D{row}'].alignment = Alignment(horizontal='right')
        ws[f'D{row}'].fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')
        
        ws[f'E{row}'] = f"₹{float(data.get('total_amount', 0)):,.2f}"
        ws[f'E{row}'].font = self.fonts['total']
        ws[f'E{row}'].alignment = Alignment(horizontal='right')
        ws[f'E{row}'].fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')
        ws[f'E{row}'].border = Border(top=Side(style='medium'), bottom=Side(style='double'))
        
        # Auto-adjust column widths based on content
        self._auto_adjust_column_widths(ws)
    
    def _build_details_sheet(self, ws, data: Dict):
        """Build detailed breakdown sheet"""
        
        row = 1
        
        # Header
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "DETAILED BREAKDOWN & METADATA"
        cell.font = self.fonts['title']
        cell.fill = PatternFill(start_color=self.colors['header'], end_color=self.colors['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 30
        row += 2
        
        # All fields (comprehensive)
        ws[f'A{row}'] = "Field"
        ws[f'B{row}'] = "Value"
        ws[f'A{row}'].font = self.fonts['subheader']
        ws[f'B{row}'].font = self.fonts['subheader']
        ws[f'A{row}'].fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')
        ws[f'B{row}'].fill = PatternFill(start_color=self.colors['total'], end_color=self.colors['total'], fill_type='solid')
        row += 1
        
        # Export all data
        for key, value in data.items():
            if key != 'line_items':  # Line items shown separately
                ws[f'A{row}'] = key.replace('_', ' ').title()
                ws[f'B{row}'] = str(value) if value is not None else 'N/A'
                row += 1
        
        # Auto-adjust column widths for details sheet
        self._auto_adjust_column_widths(ws)
    
    def _auto_adjust_column_widths(self, ws):
        """
        Automatically adjust column widths based on content length
        This ensures all content is visible and looks professional
        """
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    # Calculate length of cell value
                    if cell.value:
                        # Convert to string and get length
                        cell_value = str(cell.value)
                        
                        # Handle formulas (show result length, not formula)
                        if cell_value.startswith('='):
                            # For formulas, estimate reasonable width
                            cell_length = 12
                        else:
                            # For regular content, use actual length
                            cell_length = len(cell_value)
                        
                        # Update max length if this cell is longer
                        if cell_length > max_length:
                            max_length = cell_length
                
                except:
                    pass
            
            # Set column width with padding
            # Add 2 characters for padding, minimum width of 8
            adjusted_width = max_length + 2
            
            # Set minimum and maximum widths for usability
            if adjusted_width < 8:
                adjusted_width = 8  # Minimum width
            elif adjusted_width > 50:
                adjusted_width = 50  # Maximum width (for very long descriptions)
            
            # Apply the calculated width
            ws.column_dimensions[column_letter].width = adjusted_width


# ============ USAGE EXAMPLE ============
if __name__ == "__main__":
    # Sample invoice data
    sample_invoice = {
        'invoice_number': 'IN67/2025-26',
        'invoice_date': '2025-06-04',
        'due_date': 'N/A',
        'vendor_name': 'INNOVATION',
        'vendor_gstin': '18AABCI4851C1ZB',
        'vendor_address': 'Pattan Bazar, Guwahati, Assam',
        'currency': 'INR',
        'line_items': [
            {
                'description': 'Product A - Professional Services',
                'hsn_sac': '998314',
                'quantity': 2,
                'rate': 15000.00,
                'amount': 30000.00
            },
            {
                'description': 'Product B - Consulting',
                'hsn_sac': '998315',
                'quantity': 1,
                'rate': 10000.00,
                'amount': 10000.00
            }
        ],
        'subtotal': 33898.31,
        'cgst': 3050.85,
        'sgst': 3050.85,
        'igst': 0,
        'total_amount': 40000.00,
        'payment_status': 'unpaid',
        'created_at': '2025-10-13T10:30:00Z'
    }
    
    exporter = ProfessionalInvoiceExporter()
    exporter.export_invoice(sample_invoice, 'Professional_Invoice_Demo.xlsx')
    
    print("\n✅ Professional Excel invoice created!")
    print("Open 'Professional_Invoice_Demo.xlsx' to see the professional format")
