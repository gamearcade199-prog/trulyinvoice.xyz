"""
Clean Excel Exporter - Dynamic Column Widths (NO TEXT CLIPPING)
Single responsibility: Export invoices to properly formatted Excel files
"""

import os
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_invoices(invoices: List[Dict], filename: Optional[str] = None) -> str:
    """
    Export invoices to Excel with auto-sized columns
    
    Args:
        invoices: List of invoice dictionaries from database
        filename: Optional output filename
    
    Returns:
        str: Absolute path to Excel file
    """
    # Generate filename
    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        filename = f"invoices_{timestamp}.xlsx"
    
    # Ensure absolute path in exports directory
    if not os.path.isabs(filename):
        export_dir = os.path.join(os.getcwd(), "exports")
        os.makedirs(export_dir, exist_ok=True)
        filename = os.path.join(export_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Define columns
    columns = [
        'Invoice Number',
        'Vendor Name',
        'Invoice Date',
        'Due Date',
        'Subtotal',
        'Tax Amount',
        'Total Amount',
        'Payment Status',
        'Payment Method',
        'GSTIN',
        'Created At'
    ]
    
    # Add styled header row
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_name
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add invoice data rows
    for row_idx, invoice in enumerate(invoices, start=2):
        ws.cell(row=row_idx, column=1, value=invoice.get('invoice_number', 'N/A'))
        ws.cell(row=row_idx, column=2, value=invoice.get('vendor_name', 'N/A'))
        ws.cell(row=row_idx, column=3, value=invoice.get('invoice_date', 'N/A'))
        ws.cell(row=row_idx, column=4, value=invoice.get('due_date', 'N/A'))
        
        # Format currency values
        subtotal = invoice.get('subtotal', 0) or 0
        tax = invoice.get('tax_amount', 0) or 0
        total = invoice.get('total_amount', 0) or 0
        
        ws.cell(row=row_idx, column=5, value=f"₹{subtotal:,.2f}")
        ws.cell(row=row_idx, column=6, value=f"₹{tax:,.2f}")
        ws.cell(row=row_idx, column=7, value=f"₹{total:,.2f}")
        
        ws.cell(row=row_idx, column=8, value=invoice.get('payment_status', 'pending'))
        ws.cell(row=row_idx, column=9, value=invoice.get('payment_method', 'N/A'))
        ws.cell(row=row_idx, column=10, value=invoice.get('gstin', 'N/A'))
        ws.cell(row=row_idx, column=11, value=str(invoice.get('created_at', 'N/A')))
    
    # AUTO-SIZE ALL COLUMNS (fixes text clipping)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # Find longest content in column
        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
        
        # Set dynamic width: min 12, max 100, +3 padding
        adjusted_width = min(max(max_length + 3, 12), 100)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(filename)
    
    print(f"✅ Excel Export Complete")
    print(f"   📄 File: {os.path.basename(filename)}")
    print(f"   📊 Invoices: {len(invoices)}")
    print(f"   📏 Dynamic columns: NO TEXT CLIPPING")
    
    return filename
