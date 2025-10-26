import openpyxl
import os

# Check the Excel file structure in detail
excel_file = 'final_clip_test.xlsx'
if os.path.exists(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    print('Excel file exists:', excel_file)

    # Check the Complete Invoice Data sheet
    if 'Complete Invoice Data' in wb.sheetnames:
        ws = wb['Complete Invoice Data']
        print('Columns:', ws.max_column)
        print('Rows:', ws.max_row)

        # Check specific cells that should have long text
        print()
        print('Checking specific cells with long content:')

        # Check vendor name (should be column D, row 2)
        vendor_cell = ws.cell(row=2, column=4)  # Column D (4), Row 2
        if vendor_cell.value:
            print('Vendor Name (D2):', repr(vendor_cell.value))
            print('  Length:', len(str(vendor_cell.value)))
            print('  Wrap text:', vendor_cell.alignment.wrap_text if vendor_cell.alignment else False)
            print('  Column width:', ws.column_dimensions['D'].width)

        # Check customer name (should be column I, row 2) 
        customer_cell = ws.cell(row=2, column=9)  # Column I (9), Row 2
        if customer_cell.value:
            print()
            print('Customer Name (I2):', repr(customer_cell.value))
            print('  Length:', len(str(customer_cell.value)))
            print('  Wrap text:', customer_cell.alignment.wrap_text if customer_cell.alignment else False)
            print('  Column width:', ws.column_dimensions['I'].width)
        
        # Check address (should be column H, row 2)
        address_cell = ws.cell(row=2, column=8)  # Column H (8), Row 2
        if address_cell.value:
            print()
            print('Vendor Address (H2):', repr(address_cell.value))
            print('  Length:', len(str(address_cell.value)))
            print('  Wrap text:', address_cell.alignment.wrap_text if address_cell.alignment else False)
            print('  Column width:', ws.column_dimensions['H'].width)        # Check all cells in row 2 to see what's actually there
        print()
        print('All data in row 2:')
        for col_num in range(1, min(20, ws.max_column + 1)):  # Check first 20 columns
            cell = ws.cell(row=2, column=col_num)
            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(col_num)
                print(f'  {col_letter}{2}: {repr(cell.value)} (wrap: {cell.alignment.wrap_text if cell.alignment else False})')
        
        print()
        print('Column widths for key columns:')
        key_columns = ['D', 'K', 'N']  # Vendor Name, Vendor Address, Customer Name
        for col in key_columns:
            width = ws.column_dimensions[col].width
            print(f'  Column {col}: {width:.1f}')
            
        # Check if the text is actually fitting within the column width
        print()
        print('Text fitting analysis:')
        for col_letter, col_name in [('D', 'Vendor Name'), ('K', 'Vendor Address'), ('N', 'Customer Name')]:
            cell = ws.cell(row=2, column=openpyxl.utils.column_index_from_string(col_letter))
            if cell.value:
                text_length = len(str(cell.value))
                col_width = ws.column_dimensions[col_letter].width
                # Rough estimate: ~1.2 characters per width unit
                estimated_chars_per_line = int(col_width * 1.2)
        # Check row height
        print()
        print('Row heights:')
        for row_num in range(1, min(5, ws.max_row + 1)):
            height = ws.row_dimensions[row_num].height
            if height:
                print(f'  Row {row_num}: {height:.1f} pixels')
            else:
                print(f'  Row {row_num}: Auto (default)')