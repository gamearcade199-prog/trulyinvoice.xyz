"""
Quick Excel File Inspector - Check what's actually in the generated file
"""
from openpyxl import load_workbook
import os

excel_file = 'test_complete_invoice.xlsx'

if not os.path.exists(excel_file):
    print(f"❌ File not found: {excel_file}")
    exit(1)

print("="*80)
print(f"INSPECTING: {excel_file}")
print("="*80)

wb = load_workbook(excel_file)
print(f"\n📋 Available Sheets: {wb.sheetnames}")

for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"📊 SHEET: {sheet_name}")
    print(f"{'='*80}")
    
    ws = wb[sheet_name]
    
    # Find headers (check first 10 rows for headers)
    headers_row = None
    headers = []
    
    for row_num in range(1, min(11, ws.max_row + 1)):
        row_values = []
        for col in range(1, min(21, ws.max_column + 1)):
            value = ws.cell(row=row_num, column=col).value
            if value:
                row_values.append(str(value))
        
        # Check if this looks like a header row
        if any(keyword in ' '.join(row_values).lower() 
               for keyword in ['invoice', 'vendor', 'customer', 'amount', 'date', 'gstin', 'no', 'name']):
            headers_row = row_num
            headers = row_values
            break
    
    if headers_row:
        print(f"\n   Headers found at row {headers_row}:")
        for idx, header in enumerate(headers, 1):
            print(f"      Col {idx}: {header}")
        
        # Check for vendor GSTIN
        vendor_gstin_col = None
        for idx, header in enumerate(headers, 1):
            if 'vendor' in header.lower() and 'gstin' in header.lower():
                vendor_gstin_col = idx
                print(f"\n   ✅ Found 'Vendor GSTIN' at column {idx}")
                break
        
        if not vendor_gstin_col:
            print(f"\n   ❌ 'Vendor GSTIN' column NOT FOUND!")
            print(f"   Available columns with 'vendor': {[h for h in headers if 'vendor' in h.lower()]}")
            print(f"   Available columns with 'gstin': {[h for h in headers if 'gstin' in h.lower()]}")
        
        # Check first data row
        data_row_num = headers_row + 1
        if data_row_num <= ws.max_row:
            print(f"\n   First Data Row (row {data_row_num}):")
            for idx in range(1, min(len(headers) + 1, 11)):
                value = ws.cell(row=data_row_num, column=idx).value
                header = headers[idx-1] if idx <= len(headers) else f"Col{idx}"
                print(f"      {header}: {value}")
            
            # Specifically check vendor GSTIN value
            if vendor_gstin_col:
                gstin_value = ws.cell(row=data_row_num, column=vendor_gstin_col).value
                if gstin_value and gstin_value != '':
                    print(f"\n   ✅ Vendor GSTIN has value: {gstin_value}")
                else:
                    print(f"\n   ❌ Vendor GSTIN is EMPTY!")
    else:
        print(f"\n   No clear header row found. First 5 rows:")
        for row_num in range(1, min(6, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(6, ws.max_column + 1)):
                value = ws.cell(row=row_num, column=col).value
                row_data.append(str(value) if value else '')
            print(f"      Row {row_num}: {' | '.join(row_data)}")

wb.close()

print("\n" + "="*80)
print("✅ INSPECTION COMPLETE")
print("="*80)
