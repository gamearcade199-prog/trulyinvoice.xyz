"""
DYNAMIC EXCEL EXPORTER v1.0
Creates Excel files with columns for ANY extracted invoice data
No predefined field limitations - truly dynamic column generation
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
import re
import hashlib
from decimal import Decimal, ROUND_HALF_UP


class DynamicExcelExporter:
    """
    Dynamic Excel exporter that creates columns for ANY extracted data
    - Analyzes all available data dynamically
    - Creates appropriate columns for any field type
    - Handles nested JSON data structures
    - Professional formatting with intelligent column sizing
    """

    def __init__(self):
        # Professional color scheme
        self.colors = {
            'header_bg': '1F4E79',      # Dark blue
            'header_text': 'FFFFFF',    # White
            'accent': '4472C4',         # Medium blue
            'success': 'C6EFCE',        # Light green
            'warning': 'FFEB9C',        # Light yellow
            'error': 'FFC7CE',          # Light red
            'neutral': 'F2F2F2',        # Light grey
        }

        # Professional fonts
        self.header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.colors['header_text']
        )

        self.body_font = Font(
            name='Calibri',
            size=10
        )

        self.total_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=self.colors['accent']
        )

        # Borders
        self.thin_border = Border(
            left=Side(style='thin', color='B4B4B4'),
            right=Side(style='thin', color='B4B4B4'),
            top=Side(style='thin', color='B4B4B4'),
            bottom=Side(style='thin', color='B4B4B4')
        )

        self.thick_border = Border(
            left=Side(style='medium', color='4472C4'),
            right=Side(style='medium', color='4472C4'),
            top=Side(style='medium', color='4472C4'),
            bottom=Side(style='medium', color='4472C4')
        )

        # Fills
        self.header_fill = PatternFill(
            start_color=self.colors['header_bg'],
            end_color=self.colors['header_bg'],
            fill_type='solid'
        )

        self.accent_fill = PatternFill(
            start_color=self.colors['accent'],
            end_color=self.colors['accent'],
            fill_type='solid'
        )

        # Number formats
        self.currency_format = '₹#,##0.00'
        self.percentage_format = '0.00%'
        self.date_format = 'DD/MM/YYYY'
        self.integer_format = '#,##0'

    def export_dynamic_invoices(self, invoices: List[Dict], filename: str = None) -> str:
        """
        Export invoices with ALL extracted data to dynamically created Excel columns

        Args:
            invoices: List of invoice dictionaries with raw_extracted_data
            filename: Output filename

        Returns:
            Path to created Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"dynamic_invoices_{timestamp}.xlsx"

        # Validate input data
        validated_invoices = self._validate_and_clean_dynamic_invoices(invoices)

        if not validated_invoices:
            raise ValueError("No valid invoices to export")

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create dynamic sheets based on available data
        self._create_dynamic_sheets(wb, validated_invoices)

        # Add metadata sheet
        self._create_dynamic_metadata_sheet(wb, validated_invoices)

        # Apply final formatting
        self._apply_dynamic_formatting(wb)
        wb.save(filename)

        total_invoices = len(validated_invoices)
        total_columns = self._count_total_columns(validated_invoices)

        print(f"✅ Dynamic Excel export completed: {filename}")
        print(f"   📊 {total_invoices} invoices, {total_columns} dynamic columns")
        print(f"   🎨 Fully dynamic column generation")
        print(f"   📁 Size: {self._get_file_size_mb(filename)} MB")

        return filename

    def _validate_and_clean_dynamic_invoices(self, invoices: List[Dict]) -> List[Dict]:
        """Validate and clean invoices with dynamic data"""
        validated = []

        for idx, invoice in enumerate(invoices):
            try:
                cleaned = self._clean_dynamic_invoice_data(invoice)
                if cleaned:
                    validated.append(cleaned)
                else:
                    print(f"⚠️  Skipped invoice {idx + 1}: Invalid data structure")
            except Exception as e:
                print(f"❌ Error processing invoice {idx + 1}: {str(e)}")
                continue

        return validated

    def _clean_dynamic_invoice_data(self, invoice: Dict) -> Optional[Dict]:
        """Clean dynamic invoice data"""
        required_fields = ['invoice_number', 'vendor_name']

        # Check required fields
        for field in required_fields:
            if not invoice.get(field):
                return None

        cleaned = invoice.copy()

        # Ensure raw_extracted_data is a dict
        if not isinstance(cleaned.get('raw_extracted_data'), dict):
            cleaned['raw_extracted_data'] = {}

        # Clean numeric fields
        numeric_fields = ['total_amount', 'subtotal', 'cgst', 'sgst', 'igst', 'paid_amount']
        for field in numeric_fields:
            if cleaned.get(field) is not None:
                try:
                    cleaned[field] = float(cleaned[field])
                except:
                    cleaned[field] = 0.0

        return cleaned

    def _create_dynamic_sheets(self, wb: Workbook, invoices: List[Dict]):
        """Create sheets with dynamically analyzed columns"""

        # Analyze all available data to determine columns
        available_columns = self._analyze_all_available_columns(invoices)

        # Sheet 1: Complete Data (all fields flattened)
        complete_ws = wb.create_sheet("Complete Data")
        self._build_complete_data_sheet(complete_ws, invoices, available_columns)

        # Sheet 2: Structured Summary (traditional fields)
        summary_ws = wb.create_sheet("Structured Summary")
        self._build_structured_summary_sheet(summary_ws, invoices)

        # Sheet 3: Raw Extracted Data (JSON fields)
        raw_ws = wb.create_sheet("Raw Extracted Data")
        self._build_raw_data_sheet(raw_ws, invoices)

        # Sheet 4: Line Items (if available)
        if any(inv.get('line_items') for inv in invoices):
            items_ws = wb.create_sheet("Line Items")
            self._build_dynamic_line_items_sheet(items_ws, invoices)

    def _analyze_all_available_columns(self, invoices: List[Dict]) -> Dict[str, Dict]:
        """
        Analyze ALL available data across all invoices to create comprehensive column list
        This is the key to dynamic column generation
        """
        available_columns = {}

        # Define priority fields (always include if available)
        priority_fields = {
            'invoice_number': {'type': 'text', 'priority': 1},
            'invoice_date': {'type': 'date', 'priority': 1},
            'due_date': {'type': 'date', 'priority': 1},
            'vendor_name': {'type': 'text', 'priority': 1},
            'vendor_gstin': {'type': 'text', 'priority': 1},
            'customer_name': {'type': 'text', 'priority': 1},
            'customer_gstin': {'type': 'text', 'priority': 1},
            'total_amount': {'type': 'currency', 'priority': 1},
            'paid_amount': {'type': 'currency', 'priority': 1},
            'payment_status': {'type': 'text', 'priority': 1},
            'subtotal': {'type': 'currency', 'priority': 2},
            'cgst': {'type': 'currency', 'priority': 2},
            'sgst': {'type': 'currency', 'priority': 2},
            'igst': {'type': 'currency', 'priority': 2},
        }

        # Add priority fields if they exist in any invoice
        for field, info in priority_fields.items():
            has_data = any(inv.get(field) is not None for inv in invoices)
            if has_data:
                available_columns[field] = info

        # Analyze raw_extracted_data for additional fields
        raw_fields = set()
        for invoice in invoices:
            raw_data = invoice.get('raw_extracted_data', {})
            self._extract_fields_from_dict(raw_data, '', raw_fields)

        # Add raw fields with lower priority
        for field in raw_fields:
            if field not in available_columns:
                field_type = self._infer_field_type(field, invoices)
                available_columns[field] = {
                    'type': field_type,
                    'priority': 3,
                    'source': 'raw'
                }

        # Sort by priority then alphabetically
        sorted_columns = {}
        for priority in [1, 2, 3]:
            priority_cols = {k: v for k, v in available_columns.items() if v['priority'] == priority}
            for col in sorted(priority_cols.keys()):
                sorted_columns[col] = priority_cols[col]

        return sorted_columns

    def _extract_fields_from_dict(self, data: Dict, prefix: str, fields: set):
        """Recursively extract all field paths from nested dictionary"""
        for key, value in data.items():
            field_path = f"{prefix}{key}" if prefix else key

            if isinstance(value, dict):
                self._extract_fields_from_dict(value, f"{field_path}.", fields)
            elif isinstance(value, list):
                # For lists, add the field itself
                fields.add(field_path)
                # If list contains dicts, extract their fields too
                if value and isinstance(value[0], dict):
                    for item in value[:1]:  # Just check first item
                        self._extract_fields_from_dict(item, f"{field_path}.", fields)
            else:
                fields.add(field_path)

    def _infer_field_type(self, field_name: str, invoices: List[Dict]) -> str:
        """Infer field type based on field name and sample values"""
        field_lower = field_name.lower()

        # Currency indicators
        if any(keyword in field_lower for keyword in ['amount', 'total', 'price', 'rate', 'cgst', 'sgst', 'igst', 'tax', 'gst']):
            return 'currency'

        # Date indicators
        if any(keyword in field_lower for keyword in ['date', 'time']):
            return 'date'

        # Percentage indicators
        if 'percent' in field_lower or '%' in field_name:
            return 'percentage'

        # Numeric indicators
        if any(keyword in field_lower for keyword in ['quantity', 'qty', 'count', 'number']):
            return 'numeric'

        # Check sample values
        for invoice in invoices:
            value = self._get_nested_value(invoice, field_name)
            if value is not None:
                if isinstance(value, (int, float)):
                    return 'numeric'
                elif isinstance(value, str):
                    # Check if it's a date
                    if re.match(r'\d{4}-\d{2}-\d{2}', str(value)):
                        return 'date'
                    # Check if it's currency
                    if re.match(r'₹|RS\.?|INR', str(value), re.IGNORECASE):
                        return 'currency'

        return 'text'

    def _get_nested_value(self, data: Dict, field_path: str):
        """Get value from nested dictionary using dot notation"""
        keys = field_path.split('.')
        current = data

        try:
            for key in keys:
                if isinstance(current, dict):
                    current = current.get(key)
                elif isinstance(current, list) and key.isdigit():
                    current = current[int(key)]
                else:
                    return None
            return current
        except:
            return None

    def _build_complete_data_sheet(self, ws, invoices: List[Dict], available_columns: Dict[str, Dict]):
        """Build sheet with all available columns dynamically"""

        # Title
        ws['A1'] = 'COMPLETE INVOICE DATA EXPORT'
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=self.colors['header_bg'])
        ws.merge_cells('A1:Z1')

        # Report info
        ws['A3'] = f'Generated: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        ws['A4'] = f'Total Invoices: {len(invoices)}'
        ws['A5'] = f'Dynamic Columns: {len(available_columns)}'

        # Headers
        header_row = 7
        for col_idx, (field_name, field_info) in enumerate(available_columns.items(), start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = field_name.replace('_', ' ').title()
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Data rows
        for row, invoice in enumerate(invoices, header_row + 1):
            col_idx = 1
            for field_name, field_info in available_columns.items():
                cell = ws.cell(row=row, column=col_idx)

                # Get value based on source
                if field_info.get('source') == 'raw':
                    value = self._get_nested_value(invoice.get('raw_extracted_data', {}), field_name)
                else:
                    value = invoice.get(field_name)

                # Handle None values
                if value is None:
                    value = ''
                elif isinstance(value, list):
                    value = ', '.join(str(item) for item in value if item)
                elif isinstance(value, dict):
                    value = json.dumps(value)

                cell.value = value
                cell.border = self.thin_border
                cell.font = self.body_font

                # Apply formatting based on field type
                if field_info['type'] == 'currency':
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)):
                        cell.number_format = self.currency_format
                elif field_info['type'] == 'percentage':
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)):
                        cell.number_format = self.percentage_format
                elif field_info['type'] == 'numeric':
                    cell.alignment = Alignment(horizontal='right')
                    if isinstance(value, (int, float)):
                        cell.number_format = self.integer_format
                elif field_info['type'] == 'date':
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='left', wrap_text=True)

                col_idx += 1

        # Auto-adjust columns
        self._auto_adjust_dynamic_columns(ws, available_columns)

    def _build_structured_summary_sheet(self, ws, invoices: List[Dict]):
        """Build traditional structured summary sheet"""
        ws['A1'] = 'STRUCTURED INVOICE SUMMARY'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.colors['header_bg'])
        ws.merge_cells('A1:K1')

        # Headers
        headers = [
            'Invoice No', 'Date', 'Due Date', 'Vendor Name', 'Vendor GSTIN',
            'Customer Name', 'Total Amount', 'Paid Amount', 'Balance', 'Payment Status'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        for row, invoice in enumerate(invoices, 4):
            data = [
                invoice.get('invoice_number', ''),
                self._format_date(invoice.get('invoice_date', '')),
                self._format_date(invoice.get('due_date', '')),
                invoice.get('vendor_name', ''),
                invoice.get('vendor_gstin', ''),
                invoice.get('customer_name', ''),
                invoice.get('total_amount', 0),
                invoice.get('paid_amount', 0),
                self._calculate_balance(invoice),
                invoice.get('payment_status', 'Unpaid')
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = self.thin_border
                cell.font = self.body_font

                if col in [7, 8, 9]:
                    cell.number_format = self.currency_format
                    cell.alignment = Alignment(horizontal='right')
                elif col == 10:
                    cell.alignment = Alignment(horizontal='center')

        self._auto_adjust_columns(ws, headers)

    def _build_raw_data_sheet(self, ws, invoices: List[Dict]):
        """Build sheet showing raw extracted JSON data"""
        ws['A1'] = 'RAW EXTRACTED DATA (JSON)'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.colors['header_bg'])
        ws.merge_cells('A1:C1')

        # Headers
        headers = ['Invoice No', 'Field Name', 'Field Value']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border

        # Data - flatten JSON data
        row = 4
        for invoice in invoices:
            invoice_no = invoice.get('invoice_number', '')
            raw_data = invoice.get('raw_extracted_data', {})

            # Flatten the JSON
            flattened = self._flatten_dict(raw_data, '')

            for field_path, value in flattened.items():
                ws.cell(row=row, column=1).value = invoice_no
                ws.cell(row=row, column=2).value = field_path
                ws.cell(row=row, column=3).value = str(value) if value is not None else ''

                # Formatting
                ws.cell(row=row, column=1).border = self.thin_border
                ws.cell(row=row, column=2).border = self.thin_border
                ws.cell(row=row, column=3).border = self.thin_border
                ws.cell(row=row, column=3).alignment = Alignment(horizontal='left', wrap_text=True)

                row += 1

        self._auto_adjust_columns(ws, headers)

    def _build_dynamic_line_items_sheet(self, ws, invoices: List[Dict]):
        """Build dynamic line items sheet"""
        ws['A1'] = 'LINE ITEMS (DYNAMIC)'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.colors['header_bg'])
        ws.merge_cells('A1:Z1')

        # Analyze line item structure
        line_item_columns = self._analyze_line_item_columns(invoices)

        # Headers
        header_row = 3
        ws.cell(row=header_row, column=1).value = 'Invoice No'
        ws.cell(row=header_row, column=1).font = self.header_font
        ws.cell(row=header_row, column=1).fill = self.header_fill
        ws.cell(row=header_row, column=1).border = self.thin_border

        for col_idx, col_name in enumerate(line_item_columns, start=2):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = col_name.replace('_', ' ').title()
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Data
        row = 4
        for invoice in invoices:
            invoice_no = invoice.get('invoice_number', '')
            line_items = invoice.get('line_items', [])

            if not line_items:
                continue

            for item in line_items:
                ws.cell(row=row, column=1).value = invoice_no
                ws.cell(row=row, column=1).border = self.thin_border

                for col_idx, col_name in enumerate(line_item_columns, start=2):
                    value = item.get(col_name, '')
                    cell = ws.cell(row=row, column=col_idx)
                    cell.value = value
                    cell.border = self.thin_border

                    # Formatting
                    if col_name in ['amount', 'rate', 'quantity']:
                        cell.alignment = Alignment(horizontal='right')
                        try:
                            if col_name != 'quantity':
                                cell.number_format = self.currency_format
                        except:
                            pass

                row += 1

        # Auto adjust
        all_headers = ['Invoice No'] + line_item_columns
        self._auto_adjust_columns(ws, all_headers)

    def _analyze_line_item_columns(self, invoices: List[Dict]) -> List[str]:
        """Analyze available line item columns"""
        columns = set()

        for invoice in invoices:
            line_items = invoice.get('line_items', [])
            for item in line_items:
                if isinstance(item, dict):
                    columns.update(item.keys())

        # Priority order
        priority_cols = ['description', 'hsn_sac', 'quantity', 'unit', 'rate', 'amount']
        other_cols = sorted(columns - set(priority_cols))

        return priority_cols + other_cols

    def _create_dynamic_metadata_sheet(self, wb: Workbook, invoices: List[Dict]):
        """Create metadata sheet with dynamic export information"""
        ws = wb.create_sheet("Export Metadata")

        ws['A1'] = 'DYNAMIC EXPORT METADATA & VALIDATION'
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color=self.colors['header_bg'])
        ws.merge_cells('A1:D1')

        # Export details
        metadata = [
            ('Export Date', datetime.now().strftime('%d/%m/%Y %H:%M:%S')),
            ('Template', 'Dynamic Complete'),
            ('Total Invoices', len(invoices)),
            ('Dynamic Columns Generated', self._count_total_columns(invoices)),
            ('Total Raw Fields Extracted', self._count_raw_fields(invoices)),
            ('Exporter Version', '1.0.0 Dynamic'),
            ('Data Source', 'Complete Invoice Extraction'),
        ]

        for row, (label, value) in enumerate(metadata, 3):
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=2).value = value

        # Data quality summary
        ws['A15'] = 'DATA QUALITY SUMMARY'
        ws['A15'].font = Font(name='Calibri', size=12, bold=True)

        quality_checks = [
            ('Invoices with Raw Data', self._count_invoices_with_raw_data(invoices)),
            ('Average Fields per Invoice', self._calculate_avg_fields_per_invoice(invoices)),
            ('Structured Fields Coverage', self._calculate_structured_coverage(invoices)),
            ('Dynamic Fields Coverage', self._calculate_dynamic_coverage(invoices)),
        ]

        for row, (check, result) in enumerate(quality_checks, 16):
            ws.cell(row=row, column=1).value = check
            ws.cell(row=row, column=2).value = result

    def _apply_dynamic_formatting(self, wb: Workbook):
        """Apply final formatting to dynamic workbook"""
        for ws in wb.worksheets:
            # Add conditional formatting for payment status
            if 'Payment Status' in [cell.value for cell in ws[1] if cell.value]:
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == 'Payment Status':
                        payment_col = get_column_letter(col)
                        try:
                            paid_rule = CellIsRule(
                                operator='equal',
                                formula=['"Paid"'],
                                fill=PatternFill(start_color=self.colors['success'], fill_type='solid')
                            )
                            ws.conditional_formatting.add(f'{payment_col}2:{payment_col}1000', paid_rule)
                        except:
                            pass
                        break

            # Freeze headers
            if ws.max_row > 1:
                ws.freeze_panes = 'A2'

            # Add filters
            if ws.max_row > 1 and ws.max_column > 1:
                ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}1'

    # Utility methods
    def _flatten_dict(self, d: Dict, prefix: str = '') -> Dict[str, Any]:
        """Flatten nested dictionary"""
        flattened = {}
        for key, value in d.items():
            new_key = f"{prefix}{key}" if prefix else key
            if isinstance(value, dict):
                flattened.update(self._flatten_dict(value, f"{new_key}."))
            else:
                flattened[new_key] = value
        return flattened

    def _format_date(self, date_str: str) -> str:
        """Format date string consistently"""
        if not date_str:
            return ''
        try:
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(str(date_str), fmt)
                    return dt.strftime('%d/%m/%Y')
                except:
                    continue
            return str(date_str)
        except:
            return str(date_str)

    def _calculate_balance(self, invoice: Dict) -> float:
        """Calculate outstanding balance"""
        total = invoice.get('total_amount', 0)
        paid = invoice.get('paid_amount', 0)
        return total - paid

    def _auto_adjust_dynamic_columns(self, ws, available_columns: Dict[str, Dict]):
        """Auto-adjust columns for dynamic sheet"""
        col_idx = 1
        for field_name, field_info in available_columns.items():
            width = self._calculate_dynamic_column_width(field_name, field_info)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = width
            col_idx += 1

    def _calculate_dynamic_column_width(self, field_name: str, field_info: Dict) -> float:
        """Calculate appropriate width for dynamic column"""
        base_width = len(field_name.replace('_', ' ')) + 2

        # Type-based adjustments
        if field_info['type'] == 'currency':
            return max(base_width, 12)
        elif field_info['type'] == 'date':
            return max(base_width, 12)
        elif field_info['type'] in ['text', 'description']:
            return min(max(base_width, 20), 40)  # Reasonable limits
        else:
            return min(max(base_width, 15), 30)

    def _auto_adjust_columns(self, ws, headers: List[str]):
        """Auto-adjust column widths based on content"""
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)

            for row in range(1, min(ws.max_row + 1, 101)):
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = width

    def _get_file_size_mb(self, filename: str) -> float:
        """Get file size in MB"""
        try:
            import os
            size_bytes = os.path.getsize(filename)
            return round(size_bytes / (1024 * 1024), 2)
        except:
            return 0.0

    def _count_total_columns(self, invoices: List[Dict]) -> int:
        """Count total dynamic columns generated"""
        return len(self._analyze_all_available_columns(invoices))

    def _count_raw_fields(self, invoices: List[Dict]) -> int:
        """Count total raw fields extracted"""
        raw_fields = set()
        for invoice in invoices:
            raw_data = invoice.get('raw_extracted_data', {})
            self._extract_fields_from_dict(raw_data, '', raw_fields)
        return len(raw_fields)

    def _count_invoices_with_raw_data(self, invoices: List[Dict]) -> str:
        """Count invoices with raw extracted data"""
        with_raw = sum(1 for inv in invoices if inv.get('raw_extracted_data'))
        return f"{with_raw}/{len(invoices)}"

    def _calculate_avg_fields_per_invoice(self, invoices: List[Dict]) -> float:
        """Calculate average fields per invoice"""
        total_fields = 0
        for invoice in invoices:
            structured_fields = len([k for k, v in invoice.items() if v is not None and k != 'raw_extracted_data'])
            raw_fields = len(invoice.get('raw_extracted_data', {}))
            total_fields += structured_fields + raw_fields
        return round(total_fields / len(invoices), 1) if invoices else 0

    def _calculate_structured_coverage(self, invoices: List[Dict]) -> str:
        """Calculate structured field coverage"""
        required_fields = ['invoice_number', 'vendor_name', 'total_amount']
        covered = sum(1 for inv in invoices if all(inv.get(field) for field in required_fields))
        return f"{covered}/{len(invoices)} complete"

    def _calculate_dynamic_coverage(self, invoices: List[Dict]) -> str:
        """Calculate dynamic field coverage"""
        with_dynamic = sum(1 for inv in invoices if inv.get('raw_extracted_data'))
        return f"{with_dynamic}/{len(invoices)} with dynamic data"


# ============ USAGE EXAMPLE ============
if __name__ == "__main__":
    # Sample dynamic invoice data with ALL extracted fields
    sample_dynamic_invoices = [
        {
            'invoice_number': 'INV-2024-001',
            'invoice_date': '2024-10-15',
            'due_date': '2024-11-14',
            'vendor_name': 'ABC Technologies Pvt Ltd',
            'vendor_gstin': '29ABCDE1234F1Z5',
            'customer_name': 'XYZ Enterprises Ltd',
            'customer_gstin': '27XYZAB5678C2D3',
            'total_amount': 124154.85,
            'paid_amount': 0,
            'payment_status': 'unpaid',
            'subtotal': 103500,
            'cgst': 9315,
            'sgst': 9315,
            'igst': 0,
            'raw_extracted_data': {
                'vendor_pan': 'ABCDE1234F',
                'vendor_phone': '+91-9876543210',
                'vendor_email': 'accounts@abctech.com',
                'vendor_website': 'www.abctech.com',
                'vendor_address': '123 Business Park, Tech City, Bangalore - 560001, Karnataka, India',
                'customer_address': '456 Corporate Avenue, Mumbai - 400001, Maharashtra, India',
                'customer_phone': '+91-9876543211',
                'po_number': 'PO-2024-045',
                'challan_number': 'CH-789456',
                'eway_bill_number': '221012345678',
                'lr_number': 'LR/2024/0123',
                'shipping_charges': 2500,
                'handling_charges': 500,
                'round_off': -0.15,
                'payment_terms': '30 days from invoice date',
                'payment_method': 'Bank Transfer',
                'bank_details': 'HDFC Bank, A/c No: 1234567890, IFSC: HDFC0001234',
                'monetary_amounts': [
                    {'amount': 124154.85, 'context': 'Grand Total: Rs. 124154.85'},
                    {'amount': 103500, 'context': 'Subtotal: Rs. 103500'},
                    {'amount': 9315, 'context': 'CGST @ 9%: Rs. 9315'},
                ],
                'tax_information': {
                    'gst_numbers': ['29ABCDE1234F1Z5', '27XYZAB5678C2D3'],
                    'pan_numbers': ['ABCDE1234F'],
                    'hsn_codes': ['8471'],
                    'cess': 1035
                },
                'reference_numbers': {
                    'invoice_numbers': ['INV-2024-001'],
                    'po_numbers': ['PO-2024-045'],
                    'challan_numbers': ['CH-789456'],
                    'eway_bill_numbers': ['221012345678']
                },
                'addresses': [
                    '123 Business Park, Tech City, Bangalore - 560001, Karnataka, India',
                    '456 Corporate Avenue, Mumbai - 400001, Maharashtra, India'
                ],
                'phones': ['+91-9876543210', '+91-9876543211'],
                'emails': ['accounts@abctech.com'],
                'websites': ['www.abctech.com'],
                'date_information': [
                    {'date': '15/10/2024', 'context': 'Invoice Date: 15/10/2024'},
                    {'date': '14/11/2024', 'context': 'Due Date: 14/11/2024'}
                ],
                'text_blocks': [
                    'All disputes subject to Bangalore jurisdiction',
                    'Interest @ 24% p.a. on delayed payments',
                    'Goods once sold will not be taken back'
                ]
            },
            'line_items': [
                {
                    'description': 'Dell Latitude Laptop',
                    'hsn_sac': '8471',
                    'quantity': 2,
                    'unit': 'Pcs',
                    'rate': 45000,
                    'amount': 90000
                },
                {
                    'description': 'HP Wireless Mouse',
                    'hsn_sac': '8471',
                    'quantity': 5,
                    'unit': 'Pcs',
                    'rate': 1200,
                    'amount': 6000
                },
                {
                    'description': 'Logitech Keyboard',
                    'hsn_sac': '8471',
                    'quantity': 3,
                    'unit': 'Pcs',
                    'rate': 2500,
                    'amount': 7500
                }
            ],
            'metadata': {
                'extraction_timestamp': '2024-10-15T10:30:00',
                'content_length': 2500,
                'structured_fields_extracted': 15,
                'raw_fields_extracted': 25,
                'confidence_score': 0.95,
                'extraction_method': 'dynamic_comprehensive_v1'
            }
        }
    ]

    exporter = DynamicExcelExporter()
    result = exporter.export_dynamic_invoices(sample_dynamic_invoices, 'dynamic_invoice_export.xlsx')

    print("\n✅ Dynamic Excel export completed!")
    print("Features:")
    print("  - ✅ Complete data extraction (ALL invoice fields)")
    print("  - ✅ Dynamic column generation (no field limits)")
    print("  - ✅ Multiple sheets: Complete Data, Structured Summary, Raw Data, Line Items")
    print("  - ✅ Professional formatting with intelligent column sizing")
    print("  - ✅ Handles nested JSON data structures")
    print(f"  - 📁 File: {result}")
    print("\n🎯 This system now extracts and exports ALL invoice data!")
    print("   No more predefined field limitations - truly dynamic!")